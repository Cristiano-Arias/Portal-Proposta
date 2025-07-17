# backend_propostas.py
# Servidor Flask para sistema de propostas com melhorias

from flask import Flask, request, jsonify
from flask_cors import CORS
import smtplib
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
from email.utils import formataddr
import json
from datetime import datetime
import sqlite3
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import logging

app = Flask(__name__)
CORS(app)

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configurações de e-mail
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME', 'portaldofornecedor.arias@gmail.com')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD', 'sua_senha_app')

# Inicializar banco de dados
def init_db():
    conn = sqlite3.connect('propostas.db')
    cursor = conn.cursor()
    
    # Tabela para controle de propostas por CNPJ
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS propostas_enviadas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo TEXT NOT NULL,
            cnpj TEXT NOT NULL,
            empresa TEXT NOT NULL,
            data_envio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            protocolo TEXT NOT NULL,
            UNIQUE(processo, cnpj)
        )
    ''')
    
    conn.commit()
    conn.close()

# Verificar se CNPJ já enviou proposta para o processo
def verificar_cnpj_processo(processo, cnpj):
    conn = sqlite3.connect('propostas.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT COUNT(*) FROM propostas_enviadas 
        WHERE processo = ? AND cnpj = ?
    ''', (processo, cnpj))
    
    count = cursor.fetchone()[0]
    conn.close()
    
    return count > 0

# Registrar proposta enviada
def registrar_proposta(processo, cnpj, empresa, protocolo):
    conn = sqlite3.connect('propostas.db')
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            INSERT INTO propostas_enviadas (processo, cnpj, empresa, protocolo)
            VALUES (?, ?, ?, ?)
        ''', (processo, cnpj, empresa, protocolo))
        
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def gerar_pdf_proposta_tecnica(dados):
    """Gera PDF completo da proposta técnica com todas as informações"""
    filename = f"proposta_tecnica_{dados.get('processo', 'sem_processo')}.pdf"
    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Estilo personalizado
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1,  # Centralizado
        textColor=colors.darkblue
    )
    
    subtitulo_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.darkblue
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6,
        alignment=4  # Justificado
    )
    
    # Título principal
    story.append(Paragraph("PROPOSTA TÉCNICA", titulo_style))
    story.append(Spacer(1, 20))
    
    # Dados do processo
    story.append(Paragraph("1. DADOS DO PROCESSO", subtitulo_style))
    processo_data = [
        ['Processo:', dados.get('processo', 'N/A')],
        ['Modalidade:', dados.get('modalidade', 'N/A')],
        ['Objeto:', Paragraph(dados.get('objeto', 'N/A'), normal_style)],
        ['Valor Estimado:', f"R$ {dados.get('valor_estimado', '0'):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Prazo de Execução:', f"{dados.get('prazo_execucao', 'N/A')} dias"]
    ]
    
    table = Table(processo_data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(table)
    story.append(Spacer(1, 20))
    
    # Dados da empresa
    story.append(Paragraph("2. DADOS DA EMPRESA", subtitulo_style))
    empresa_data = [
        ['Razão Social:', dados.get('razao_social', 'N/A')],
        ['CNPJ:', dados.get('cnpj', 'N/A')],
        ['Endereço:', dados.get('endereco', 'N/A')],
        ['Telefone:', dados.get('telefone', 'N/A')],
        ['E-mail:', dados.get('email', 'N/A')],
        ['Responsável Técnico:', dados.get('responsavel_tecnico', 'N/A')]
    ]
    
    table = Table(empresa_data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(table)
    story.append(Spacer(1, 20))
    
    # Proposta técnica detalhada
    story.append(Paragraph("3. PROPOSTA TÉCNICA DETALHADA", subtitulo_style))
    
    # 3.1 Objeto da licitação
    story.append(Paragraph("3.1 Objeto da Licitação", styles['Heading3']))
    story.append(Paragraph(dados.get('objeto_licitacao', 'Não informado'), normal_style))
    story.append(Spacer(1, 12))
    
    # 3.2 Escopo dos serviços
    story.append(Paragraph("3.2 Descrição Detalhada do Escopo", styles['Heading3']))
    story.append(Paragraph("<b>Serviços Inclusos:</b>", normal_style))
    story.append(Paragraph(dados.get('escopo_inclusos', 'Não informado'), normal_style))
    story.append(Spacer(1, 6))
    story.append(Paragraph("<b>Serviços Não Inclusos:</b>", normal_style))
    story.append(Paragraph(dados.get('escopo_exclusos', 'Não informado'), normal_style))
    story.append(Spacer(1, 12))
    
    # 3.3 Metodologia
    story.append(Paragraph("3.3 Metodologia de Execução", styles['Heading3']))
    story.append(Paragraph(dados.get('metodologia', 'Não informado'), normal_style))
    story.append(Spacer(1, 6))
    story.append(Paragraph("<b>Plano de Execução:</b>", normal_style))
    story.append(Paragraph(dados.get('sequencia_execucao', 'Não informado'), normal_style))
    story.append(Spacer(1, 12))
    
    # 3.3.1 Tabela de Serviços (sem valores)
    if dados.get('servicos_tecnica'):
        story.append(Paragraph("3.3.1 Tabela de Serviços", styles['Heading3']))
        servicos_data = [['Item', 'Descrição', 'Unid', 'Quantidade']]
        for servico in dados.get('servicos_tecnica', []):
            servicos_data.append([
                servico[0],  # Item
                servico[1],  # Descrição
                servico[2],  # Unidade
                servico[3]   # Quantidade
            ])
        
        if len(servicos_data) > 1:
            table = Table(servicos_data, colWidths=[0.8*inch, 3*inch, 0.8*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
    
    # 3.4 Mão de Obra
    story.append(Paragraph("3.4 Mão de Obra Direta e Indireta", styles['Heading3']))
    if dados.get('equipe_tecnica'):
        equipe_data = [['Função', 'Quantidade', 'Tempo (meses)']]
        for func in dados.get('equipe_tecnica', []):
            equipe_data.append([func[0], func[1], func[2]])
        
        if len(equipe_data) > 1:
            table = Table(equipe_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
    
    # 3.5 Lista de Materiais
    story.append(Paragraph("3.5 Lista de Materiais", styles['Heading3']))
    if dados.get('materiais'):
        materiais_data = [['Material', 'Especificação', 'Unidade', 'Quantidade']]
        for mat in dados.get('materiais', []):
            materiais_data.append([mat[0], mat[1], mat[2], mat[3]])
        
        if len(materiais_data) > 1:
            table = Table(materiais_data, colWidths=[1.5*inch, 2*inch, 0.8*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
    
    # 3.6 Lista de Equipamentos
    story.append(Paragraph("3.6 Lista de Equipamentos", styles['Heading3']))
    if dados.get('equipamentos'):
        equip_data = [['Equipamento', 'Especificação', 'Unid', 'Qtd', 'Tempo']]
        for eq in dados.get('equipamentos', []):
            equip_data.append([eq[0], eq[1], eq[2], eq[3], eq[4] + ' meses'])
        
        if len(equip_data) > 1:
            table = Table(equip_data, colWidths=[1.5*inch, 1.8*inch, 0.6*inch, 0.6*inch, 0.8*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
    
    # Nova página para continuação
    story.append(PageBreak())
    
    # 3.9 Prazo de Execução
    story.append(Paragraph("3.9 Prazo de Execução", styles['Heading3']))
    story.append(Paragraph(f"Prazo total de execução: <b>{dados.get('prazo_execucao', 'N/A')}</b>", normal_style))
    story.append(Spacer(1, 12))
    
    # 3.10 Prazo de Mobilização
    story.append(Paragraph("3.10 Prazo de Mobilização", styles['Heading3']))
    story.append(Paragraph(f"Prazo de mobilização: <b>{dados.get('prazo_mobilizacao', 'N/A')}</b>", normal_style))
    story.append(Spacer(1, 12))
    
    # 3.11 Cronograma
    story.append(Paragraph("3.11 Cronograma de Execução", styles['Heading3']))
    if dados.get('cronograma'):
        crono_data = [['Atividade', 'Duração', 'Início', 'Fim']]
        for ativ in dados.get('cronograma', []):
            crono_data.append([ativ[0], ativ[1], ativ[2], ativ[3]])
        
        if len(crono_data) > 1:
            table = Table(crono_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
    
    # 3.12 Garantias
    story.append(Paragraph("3.12 Garantias Oferecidas", styles['Heading3']))
    story.append(Paragraph(dados.get('garantias', 'Não informado'), normal_style))
    story.append(Spacer(1, 12))
    
    # 3.13 Necessidade de Canteiro
    story.append(Paragraph("3.13 Necessidade de Canteiro", styles['Heading3']))
    story.append(Paragraph(dados.get('estrutura_canteiro', 'Não informado'), normal_style))
    story.append(Spacer(1, 12))
    
    # 3.14 Obrigações
    story.append(Paragraph("3.14 Obrigações", styles['Heading3']))
    story.append(Paragraph("<b>Responsabilidades da CONTRATADA:</b>", normal_style))
    story.append(Paragraph(dados.get('obrigacoes_contratada', 'Não informado'), normal_style))
    story.append(Spacer(1, 6))
    story.append(Paragraph("<b>Responsabilidades da CONTRATANTE:</b>", normal_style))
    story.append(Paragraph(dados.get('obrigacoes_contratante', 'Não informado'), normal_style))
    story.append(Spacer(1, 12))
    
    # 3.15 Experiência
    story.append(Paragraph("3.15 Experiência", styles['Heading3']))
    if dados.get('experiencia'):
        exp_data = [['Obra', 'Cliente', 'Valor', 'Ano', 'Contato']]
        for exp in dados.get('experiencia', []):
            exp_data.append([exp[0], exp[1], exp[2], exp[3], exp[4]])
        
        if len(exp_data) > 1:
            table = Table(exp_data, colWidths=[1.8*inch, 1.5*inch, 1*inch, 0.7*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
    
    # 3.16 Certificações
    story.append(Paragraph("3.16 Certificações/Qualificações", styles['Heading3']))
    if dados.get('certificacoes'):
        cert_data = [['Certificação', 'Órgão', 'Validade']]
        for cert in dados.get('certificacoes', []):
            cert_data.append([cert[0], cert[1], cert[2]])
        
        if len(cert_data) > 1:
            table = Table(cert_data, colWidths=[2*inch, 2*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
    
    # Assinatura
    story.append(Spacer(1, 40))
    story.append(Paragraph("_" * 60, styles['Normal']))
    story.append(Paragraph(dados.get('razao_social', 'EMPRESA'), styles['Normal']))
    story.append(Paragraph(f"CNPJ: {dados.get('cnpj', '')}", styles['Normal']))
    story.append(Paragraph(f"Responsável: {dados.get('responsavel_tecnico', '')}", styles['Normal']))
    
    doc.build(story)
    return filename

def gerar_pdf_proposta_comercial(dados):
    """Gera PDF da proposta comercial"""
    filename = f"proposta_comercial_{dados.get('processo', 'sem_processo')}.pdf"
    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1,
        textColor=colors.darkblue
    )
    
    story.append(Paragraph("PROPOSTA COMERCIAL", titulo_style))
    story.append(Spacer(1, 20))
    
    # Dados do processo
    story.append(Paragraph("DADOS DO PROCESSO", styles['Heading2']))
    processo_info = f"Processo: {dados.get('processo', 'N/A')} | Empresa: {dados.get('razao_social', 'N/A')}"
    story.append(Paragraph(processo_info, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Planilha de custos
    story.append(Paragraph("PLANILHA DE CUSTOS", styles['Heading2']))
    
    # Cabeçalho da tabela
    custos_data = [['Item', 'Descrição', 'Quantidade', 'Valor Unitário', 'Valor Total']]
    
    # Itens da planilha
    itens = dados.get('itens_orcamento', [])
    total_geral = 0
    
    for i, item in enumerate(itens, 1):
        quantidade = float(item.get('quantidade', 0))
        valor_unitario = float(item.get('valor_unitario', 0))
        valor_total = quantidade * valor_unitario
        total_geral += valor_total
        
        custos_data.append([
            str(i),
            item.get('descricao', ''),
            f"{quantidade:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"R$ {valor_unitario:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        ])
    
    # Total
    custos_data.append(['', '', '', 'TOTAL GERAL:', f"R$ {total_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')])
    
    table = Table(custos_data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    story.append(Spacer(1, 20))
    
    # Condições comerciais
    story.append(Paragraph("CONDIÇÕES COMERCIAIS", styles['Heading2']))
    story.append(Paragraph(f"Prazo de Execução: {dados.get('prazo_execucao', 'N/A')} dias", styles['Normal']))
    story.append(Paragraph(f"Condições de Pagamento: {dados.get('condicoes_pagamento', 'N/A')}", styles['Normal']))
    story.append(Paragraph(f"Validade da Proposta: {dados.get('validade_proposta', '60')} dias", styles['Normal']))
    
    doc.build(story)
    return filename

def gerar_excel_orcamento(dados):
    """Gera planilha Excel do orçamento"""
    filename = f"orcamento_{dados.get('processo', 'sem_processo')}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orçamento"
    
    # Cabeçalho
    ws['A1'] = "ORÇAMENTO DETALHADO"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    
    ws['A3'] = f"Processo: {dados.get('processo', 'N/A')}"
    ws['A4'] = f"Empresa: {dados.get('razao_social', 'N/A')}"
    ws['A5'] = f"Data: {datetime.now().strftime('%d/%m/%Y')}"
    
    # Cabeçalho da tabela
    headers = ['Item', 'Descrição', 'Quantidade', 'Valor Unitário', 'Valor Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Dados
    itens = dados.get('itens_orcamento', [])
    total_geral = 0
    
    for i, item in enumerate(itens, 1):
        row = 7 + i
        quantidade = float(item.get('quantidade', 0))
        valor_unitario = float(item.get('valor_unitario', 0))
        valor_total = quantidade * valor_unitario
        total_geral += valor_total
        
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item.get('descricao', ''))
        ws.cell(row=row, column=3, value=quantidade)
        ws.cell(row=row, column=4, value=valor_unitario)
        ws.cell(row=row, column=5, value=valor_total)
    
    # Total
    total_row = 7 + len(itens) + 1
    ws.cell(row=total_row, column=4, value="TOTAL GERAL:").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=total_geral).font = Font(bold=True)
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    wb.save(filename)
    return filename

def enviar_email_smtp_direto(destinatario, assunto, html_content, anexos=None):
    """Envia e-mail com encoding UTF-8 correto"""
    try:
        # Configurar servidor SMTP
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.starttls()
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        
        # Criar mensagem
        msg = MIMEMultipart('alternative')
        msg['From'] = formataddr(('Sistema de Propostas', app.config['MAIL_USERNAME']))
        msg['To'] = destinatario
        msg['Subject'] = Header(assunto, 'utf-8')
        
        # Adicionar conteúdo HTML
        html_part = MIMEText(html_content, 'html', 'utf-8')
        msg.attach(html_part)
        
        # Adicionar anexos
        if anexos:
            for anexo in anexos:
                if os.path.exists(anexo):
                    with open(anexo, "rb") as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                    
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename= {os.path.basename(anexo)}'
                    )
                    msg.attach(part)
        
        # Enviar
        server.send_message(msg)
        server.quit()
        
        logger.info(f"E-mail enviado com sucesso para {destinatario}")
        return True
        
    except Exception as e:
        logger.error(f"Erro ao enviar e-mail: {str(e)}")
        return False

@app.route('/verificar-cnpj', methods=['POST'])
def verificar_cnpj():
    """Verifica se CNPJ já enviou proposta para o processo"""
    data = request.json
    processo = data.get('processo')
    cnpj = data.get('cnpj')
    
    if not processo or not cnpj:
        return jsonify({'error': 'Processo e CNPJ são obrigatórios'}), 400
    
    ja_enviou = verificar_cnpj_processo(processo, cnpj)
    
    return jsonify({
        'ja_enviou': ja_enviou,
        'mensagem': 'CNPJ já enviou proposta para este processo' if ja_enviou else 'CNPJ pode enviar proposta'
    })

@app.route('/enviar-proposta', methods=['POST'])
def enviar_proposta():
    """Endpoint principal para envio de propostas"""
    try:
        data = request.json
        logger.info(f"Dados recebidos: {json.dumps(data, indent=2)}")
        
        # Validações básicas
        processo = data.get('processo')
        cnpj = data.get('cnpj')
        empresa = data.get('razao_social')
        
        if not all([processo, cnpj, empresa]):
            return jsonify({'error': 'Dados obrigatórios não fornecidos'}), 400
        
        # Verificar se CNPJ já enviou proposta
        if verificar_cnpj_processo(processo, cnpj):
            return jsonify({
                'error': 'Este CNPJ já enviou uma proposta para este processo',
                'codigo': 'CNPJ_JA_ENVIADO'
            }), 409
        
        # Gerar protocolo
        protocolo = f"PROP-{datetime.now().strftime('%Y%m%d-%H%M-%S')}"
        
        # Preparar dados completos para os PDFs
        dados_completos = {
            'processo': processo,
            'protocolo': protocolo,
            'razao_social': empresa,
            'cnpj': cnpj,
            'endereco': data.get('endereco', ''),
            'telefone': data.get('telefone', ''),
            'email': data.get('email', ''),
            'responsavel_tecnico': data.get('responsavel_tecnico', ''),
            'modalidade': data.get('modalidade', 'Pregão'),
            'objeto': data.get('objeto', ''),
            'valor_estimado': float(data.get('valor_estimado', 0)),
            'prazo_execucao': data.get('prazo_execucao', ''),
            
            # Dados técnicos
            'objeto_licitacao': data.get('objeto_licitacao', ''),
            'escopo_inclusos': data.get('escopo_inclusos', ''),
            'escopo_exclusos': data.get('escopo_exclusos', ''),
            'metodologia': data.get('metodologia', ''),
            'sequencia_execucao': data.get('sequencia_execucao', ''),
            'servicos_tecnica': data.get('servicos_tecnica', []),
            'equipe_tecnica': data.get('equipe_tecnica', []),
            'materiais': data.get('materiais', []),
            'equipamentos': data.get('equipamentos', []),
            'prazo_mobilizacao': data.get('prazo_mobilizacao', ''),
            'cronograma': data.get('cronograma', []),
            'garantias': data.get('garantias', ''),
            'estrutura_canteiro': data.get('estrutura_canteiro', ''),
            'obrigacoes_contratada': data.get('obrigacoes_contratada', ''),
            'obrigacoes_contratante': data.get('obrigacoes_contratante', ''),
            'experiencia': data.get('experiencia', []),
            'certificacoes': data.get('certificacoes', []),
            
            # Dados comerciais
            'valor_total_proposta': data.get('valor_total', '0'),
            'condicoes_pagamento': data.get('condicoes_pagamento', ''),
            'validade_proposta': data.get('validade_proposta', '60'),
            'itens_orcamento': data.get('itens_orcamento', [])
        }
        
        # Gerar anexos
        pdf_tecnica = gerar_pdf_proposta_tecnica(dados_completos)
        pdf_comercial = gerar_pdf_proposta_comercial(dados_completos)
        excel_orcamento = gerar_excel_orcamento(dados_completos)
        
        # Preparar e-mail
        email_suprimentos = os.getenv('EMAIL_SUPRIMENTOS', 'portaldofornecedor.arias@gmail.com')
        assunto = f"Nova Proposta - Processo {processo} - {empresa}"
        
        # Template de e-mail melhorado
        html_content = f"""
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Nova Proposta Recebida</title>
        </head>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 800px; margin: 0 auto; padding: 20px;">
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; text-align: center; margin-bottom: 30px;">
                <h1 style="margin: 0; font-size: 28px;">📋 Nova Proposta Recebida</h1>
                <p style="margin: 10px 0 0 0; font-size: 16px; opacity: 0.9;">Sistema de Gestão de Propostas</p>
            </div>
            
            <div style="background: #f8f9fa; padding: 25px; border-radius: 8px; margin-bottom: 25px;">
                <h2 style="color: #495057; margin-top: 0; border-bottom: 2px solid #dee2e6; padding-bottom: 10px;">🏢 Dados da Empresa</h2>
                <div style="display: grid; gap: 10px;">
                    <p><strong>Razão Social:</strong> {dados_completos.get('razao_social', 'N/A')}</p>
                    <p><strong>CNPJ:</strong> {dados_completos.get('cnpj', 'N/A')}</p>
                    <p><strong>Endereço:</strong> {dados_completos.get('endereco', 'N/A')}</p>
                    <p><strong>Telefone:</strong> {dados_completos.get('telefone', 'N/A')}</p>
                    <p><strong>E-mail:</strong> {dados_completos.get('email', 'N/A')}</p>
                    <p><strong>Responsável Técnico:</strong> {dados_completos.get('responsavel_tecnico', 'N/A')}</p>
                </div>
            </div>
            
            <div style="background: #e3f2fd; padding: 25px; border-radius: 8px; margin-bottom: 25px;">
                <h2 style="color: #1976d2; margin-top: 0; border-bottom: 2px solid #bbdefb; padding-bottom: 10px;">📊 Dados do Processo</h2>
                <div style="display: grid; gap: 10px;">
                    <p><strong>Número do Processo:</strong> {processo}</p>
                    <p><strong>Modalidade:</strong> {dados_completos.get('modalidade', 'N/A')}</p>
                    <p><strong>Objeto:</strong> {dados_completos.get('objeto', 'N/A')}</p>
                    <p><strong>Valor Estimado:</strong> R$ {float(dados_completos.get('valor_estimado', 0)):,.2f}</p>
                    <p><strong>Protocolo:</strong> {protocolo}</p>
                    <p><strong>Data/Hora:</strong> {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}</p>
                </div>
            </div>
            
            <div style="background: #e8f5e8; padding: 25px; border-radius: 8px; margin-bottom: 25px;">
                <h2 style="color: #2e7d32; margin-top: 0; border-bottom: 2px solid #c8e6c9; padding-bottom: 10px;">💰 Resumo Comercial</h2>
                <div style="display: grid; gap: 10px;">
                    <p><strong>Valor Total da Proposta:</strong> R$ {dados_completos.get('valor_total_proposta', '0,00')}</p>
                    <p><strong>Prazo de Execução:</strong> {dados_completos.get('prazo_execucao', 'N/A')} dias</p>
                    <p><strong>Condições de Pagamento:</strong> {dados_completos.get('condicoes_pagamento', 'N/A')}</p>
                    <p><strong>Validade da Proposta:</strong> {dados_completos.get('validade_proposta', '60')} dias</p>
                </div>
            </div>
            
            <div style="background: #fff3e0; padding: 25px; border-radius: 8px; margin-bottom: 25px;">
                <h2 style="color: #f57c00; margin-top: 0; border-bottom: 2px solid #ffcc02; padding-bottom: 10px;">📎 Anexos Inclusos</h2>
                <ul style="list-style-type: none; padding: 0;">
                    <li style="padding: 8px 0; border-bottom: 1px solid #ffcc02;">📄 Proposta Técnica Completa (PDF)</li>
                    <li style="padding: 8px 0; border-bottom: 1px solid #ffcc02;">💼 Proposta Comercial (PDF)</li>
                    <li style="padding: 8px 0;">📊 Planilha de Orçamento (Excel)</li>
                </ul>
            </div>
            
            <div style="background: #f3e5f5; padding: 25px; border-radius: 8px; margin-bottom: 25px;">
                <h2 style="color: #7b1fa2; margin-top: 0; border-bottom: 2px solid #ce93d8; padding-bottom: 10px;">⚙️ Próximos Passos</h2>
                <ol style="padding-left: 20px;">
                    <li>Análise da documentação técnica</li>
                    <li>Verificação da proposta comercial</li>
                    <li>Avaliação dos critérios de habilitação</li>
                    <li>Comunicação do resultado</li>
                </ol>
            </div>
            
            <div style="text-align: center; padding: 20px; background: #f8f9fa; border-radius: 8px; margin-top: 30px;">
                <p style="margin: 0; color: #6c757d; font-size: 14px;">
                    Este e-mail foi gerado automaticamente pelo Sistema de Gestão de Propostas<br>
                    Para dúvidas, entre em contato com o setor de suprimentos
                </p>
            </div>
        </body>
        </html>
        """
        
        # Enviar e-mail
        anexos = [pdf_tecnica, pdf_comercial, excel_orcamento]
        email_enviado = enviar_email_smtp_direto(email_suprimentos, assunto, html_content, anexos)
        
        if email_enviado:
            # Registrar proposta no banco
            registrar_proposta(processo, cnpj, empresa, protocolo)
            
            # Limpar arquivos temporários
            for arquivo in anexos:
                if os.path.exists(arquivo):
                    os.remove(arquivo)
            
            return jsonify({
                'success': True,
                'protocolo': protocolo,
                'mensagem': 'Proposta enviada com sucesso!'
            })
        else:
            return jsonify({'error': 'Erro ao enviar e-mail'}), 500
            
    except Exception as e:
        logger.error(f"Erro no envio da proposta: {str(e)}")
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/status', methods=['GET'])
def status():
    """Endpoint de status do servidor"""
    return jsonify({
        'status': 'online',
        'timestamp': datetime.now().isoformat(),
        'version': '2.0'
    })

if __name__ == '__main__':
    init_db()
    print("🚀 Servidor iniciado com sucesso!")
    print("📧 Sistema de e-mail configurado")
    print("🗄️ Banco de dados inicializado")
    print("🔒 Controle de CNPJ ativo")
    app.run(host='0.0.0.0', port=5000, debug=True)
